"""
Load 2025RatesHistory1990-2025.xlsx into county_tax_rate.

Sheet layout: one row per taxing entity.
Columns: TDC (entity_code), JURISNAME, RATE25, RATE24, … RATE90
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import config
from loaders.db import get_conn, execute_schema, batch_upsert
from loaders.scrape_billing_history import DEFAULT_COUNTY  # DALLAS-GATE-4 / PARCEL-ROLLUP-HOTFIX-1

import openpyxl


def load(conn, county_code=DEFAULT_COUNTY):
    wb = openpyxl.load_workbook(config.TAX_RATES_XL, data_only=True)
    ws = wb.active

    # Find header row (first row with 'TDC' in col A)
    header_row = None
    for row in ws.iter_rows(values_only=True):
        if row[0] == "TDC":
            header_row = row
            break
    if not header_row:
        raise ValueError("Could not find TDC header row in tax rates XLSX")

    # Build year→col mapping from column names like RATE25, RATE24 …
    year_cols = {}
    for idx, cell in enumerate(header_row):
        if cell and str(cell).startswith("RATE") and len(str(cell)) == 6:
            try:
                yr_suffix = int(str(cell)[4:])
                year = 2000 + yr_suffix if yr_suffix <= 30 else 1900 + yr_suffix
                year_cols[idx] = year
            except ValueError:
                pass

    rows = []
    # DALLAS-GATE-4 family completion (PX-20260822-06-rev1): county_code
    # added first in the column list/VALUES/ON CONFLICT target, matching
    # the PARCEL-ROLLUP-HOTFIX-1 convention -- live PK for county_tax_rate
    # is (county_code, entity_code, tax_year), confirmed via \d against
    # production, 2026-08-23. This loader writes ONE county per invocation
    # (county_code param), so every row it produces shares the same value.
    upsert_sql = """
        INSERT INTO county_tax_rate (county_code, entity_code, entity_name, tax_year, rate)
        VALUES (%s, %s, %s, %s, %s)
        ON CONFLICT (county_code, entity_code, tax_year) DO UPDATE
            SET entity_name = EXCLUDED.entity_name,
                rate        = EXCLUDED.rate
    """

    for row in ws.iter_rows(values_only=True):
        entity_code = row[0]
        entity_name = row[1]
        if not entity_code or entity_code == "TDC":
            continue
        for col_idx, year in year_cols.items():
            val = row[col_idx]
            if val is None or val == "-" or val == "":
                continue
            try:
                rate = float(val)   # already decimal per $100 (e.g. 0.375845)
            except (ValueError, TypeError):
                continue
            rows.append((county_code, str(entity_code), str(entity_name), year, rate))

    n = batch_upsert(conn, upsert_sql, rows)
    print(f"  county_tax_rate: {n:,} rows loaded")
    return n


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--county", default=DEFAULT_COUNTY,
        help=f"county_code written to every county_tax_rate row "
             f"(default: {DEFAULT_COUNTY}). DALLAS-GATE-4 / "
             f"PARCEL-ROLLUP-HOTFIX-1 convention.",
    )
    args = ap.parse_args()

    conn = get_conn()
    execute_schema(conn)
    load(conn, county_code=args.county)
    conn.close()
